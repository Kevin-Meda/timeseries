"""Test script to generate sample data and run the pipeline."""

import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
from dateutil.relativedelta import relativedelta

def generate_test_data():
    """Generate a test Excel file with 3 categories and 60 months of data."""

    # Create date range - 60 months
    start_date = datetime(2019, 1, 1)
    dates = [start_date + relativedelta(months=i) for i in range(60)]
    date_strings = [d.strftime("%m.%Y") for d in dates]

    np.random.seed(42)

    # Category 1: Seasonal pattern with trend
    t = np.arange(60)
    cat1 = 1000 + 50 * t + 200 * np.sin(2 * np.pi * t / 12) + np.random.normal(0, 50, 60)

    # Category 2: Flat with noise
    cat2 = 500 + np.random.normal(0, 30, 60)

    # Category 3: Growing trend
    cat3 = 200 + 20 * t + np.random.normal(0, 40, 60)

    # Introduce some negative values
    cat1[15] = -50
    cat2[30] = -100
    cat3[45] = -20

    # Introduce some NaN values
    cat1[25] = np.nan
    cat2[10] = np.nan
    cat2[40] = np.nan
    cat3[55] = np.nan

    # Introduce outliers (values far from local trend)
    cat1[35] = 10000  # Big outlier - should be ~2700 based on trend
    cat3[20] = 5000   # Big outlier - should be ~600 based on trend

    # Create DataFrame
    data = {
        "Product": ["Product A", "Product B", "Product C"],
    }

    # Add date columns
    for i, date_str in enumerate(date_strings):
        data[date_str] = [cat1[i], cat2[i], cat3[i]]

    df = pd.DataFrame(data)

    # Create header row with dates
    output_path = Path("data/input/demand.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Write with proper structure: first row is dates, first column is product names
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    ws = wb.active

    # Write header row with dates as text
    ws.cell(row=1, column=1, value="")
    for col_idx, date_str in enumerate(date_strings, start=2):
        cell = ws.cell(row=1, column=col_idx, value=date_str)
        cell.number_format = "@"  # Text format

    # Write data rows
    data_rows = [
        ("Product A", cat1),
        ("Product B", cat2),
        ("Product C", cat3),
    ]

    for row_idx, (name, values) in enumerate(data_rows, start=2):
        ws.cell(row=row_idx, column=1, value=name)
        for col_idx, val in enumerate(values, start=2):
            if pd.isna(val):
                ws.cell(row=row_idx, column=col_idx, value=None)
            else:
                ws.cell(row=row_idx, column=col_idx, value=float(val))

    wb.save(output_path)

    print(f"Test data created at: {output_path}")
    print(f"  - 3 categories: Product A, Product B, Product C")
    print(f"  - 60 months of data: {date_strings[0]} to {date_strings[-1]}")
    print(f"  - Includes negative values, NaN, and outliers")

    return output_path


def run_test():
    """Run the full pipeline test."""
    import sys
    sys.path.insert(0, "src")

    # Generate test data
    print("=" * 60)
    print("STEP 1: Generating test data")
    print("=" * 60)
    generate_test_data()

    # Run the pipeline
    print("\n" + "=" * 60)
    print("STEP 2: Running forecasting pipeline")
    print("=" * 60)

    from forecast.pipeline import run_pipeline

    results = run_pipeline(
        config_dir="configs",
        log_level="INFO",
    )

    print("\n" + "=" * 60)
    print("STEP 3: Results Summary")
    print("=" * 60)

    if "error" in results:
        print(f"Pipeline failed: {results['error']}")
        return

    print(f"\nCategories processed: {results.get('categories_processed', 0)}")

    model_results = results.get("model_results", {})
    for category, models in model_results.items():
        print(f"\n{category}:")
        for model_name, metrics in models.items():
            print(f"  {model_name}:")
            print(f"    RMSE: {metrics['rmse']:.2f}")
            print(f"    MAPE: {metrics['mape']:.2%}")
            print(f"    MAE: {metrics['mae']:.2f}")

    ensemble_results = results.get("ensemble_results", {})
    print("\nEnsemble Forecasts:")
    for category, info in ensemble_results.items():
        models_used = info.get("models_used", [])
        weights = info.get("weights", {})
        print(f"  {category}: {', '.join(models_used)}")
        if weights:
            weight_str = ", ".join(f"{k}={v:.2f}" for k, v in weights.items())
            print(f"    Weights: {weight_str}")

    print("\n" + "=" * 60)
    print("Test completed successfully!")
    print("=" * 60)
    print("\nOutput files:")
    print("  - output/results/model_results.xlsx")
    print("  - output/results/ensemble_forecasts.xlsx")
    print("  - output/plots/preprocessing/")
    print("  - output/plots/evaluation/")
    print("  - output/plots/forecast/")
    print("  - output/models/")


if __name__ == "__main__":
    run_test()
