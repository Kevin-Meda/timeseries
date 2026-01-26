"""Test script with Optuna optimization and a skipped product (insufficient data)."""

import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
from dateutil.relativedelta import relativedelta
import shutil


def generate_test_data_with_sparse_product():
    """Generate test data with 4 categories, one with insufficient data."""

    # Create date range - 60 months
    start_date = datetime(2019, 1, 1)
    dates = [start_date + relativedelta(months=i) for i in range(60)]
    date_strings = [d.strftime("%m.%Y") for d in dates]

    np.random.seed(42)

    # Category 1: Seasonal pattern with trend (full data)
    t = np.arange(60)
    cat1 = 1000 + 50 * t + 200 * np.sin(2 * np.pi * t / 12) + np.random.normal(0, 50, 60)

    # Category 2: Flat with noise (full data)
    cat2 = 500 + np.random.normal(0, 30, 60)

    # Category 3: Growing trend (full data)
    cat3 = 200 + 20 * t + np.random.normal(0, 40, 60)

    # Category 4: Sparse/Insufficient data - only last 6 months have data
    # This should be skipped due to insufficient lookback data
    cat4 = [0.0] * 54 + list(100 + np.random.normal(0, 10, 6))

    # Add some noise to full categories
    cat1[15] = -50
    cat2[30] = -100
    cat3[45] = -20

    cat1[25] = np.nan
    cat2[10] = np.nan
    cat3[55] = np.nan

    cat1[35] = 10000
    cat3[20] = 5000

    output_path = Path("data/input/demand_optuna.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    # Write header row with dates as text
    ws.cell(row=1, column=1, value="")
    for col_idx, date_str in enumerate(date_strings, start=2):
        cell = ws.cell(row=1, column=col_idx, value=date_str)
        cell.number_format = "@"

    # Write data rows - 4 products, one sparse
    data_rows = [
        ("Product A", cat1),
        ("Product B", cat2),
        ("Product C", cat3),
        ("Product D (Sparse)", cat4),  # This one has insufficient data
    ]

    for row_idx, (name, values) in enumerate(data_rows, start=2):
        ws.cell(row=row_idx, column=1, value=name)
        for col_idx, val in enumerate(values, start=2):
            if pd.isna(val) or val == 0:
                ws.cell(row=row_idx, column=col_idx, value=None if pd.isna(val) else 0)
            else:
                ws.cell(row=row_idx, column=col_idx, value=float(val))

    wb.save(output_path)

    print(f"Test data created at: {output_path}")
    print(f"  - 4 categories: Product A, B, C (full data), Product D (sparse)")
    print(f"  - 60 months of data: {date_strings[0]} to {date_strings[-1]}")
    print(f"  - Product D has only 6 months of non-zero data (should be skipped)")

    return output_path


def create_optuna_config():
    """Create temporary config with Optuna enabled."""
    config_dir = Path("configs_optuna")
    config_dir.mkdir(parents=True, exist_ok=True)

    # Copy base configs
    for config_file in Path("configs").glob("*.yaml"):
        shutil.copy(config_file, config_dir / config_file.name)

    # Update models.yaml to enable Optuna
    models_config = """# Model configuration with Optuna optimization
sarima:
  enabled: true
  use_optuna: true
  optuna_trials: 20  # Fewer trials for testing

chronos:
  enabled: true

holt_winters:
  enabled: true
"""
    with open(config_dir / "models.yaml", "w") as f:
        f.write(models_config)

    # Update preprocessing to be stricter for sparse data detection
    preprocess_config = """# Preprocessing configuration
classification:
  lookback_months: 24
  min_nonzero_pct: 0.5  # At least 50% non-zero in lookback

data_window_months: 60
outlier_std_threshold: 3.0
"""
    with open(config_dir / "preprocessing.yaml", "w") as f:
        f.write(preprocess_config)

    print(f"Optuna config created at: {config_dir}")
    return str(config_dir)


def run_optuna_test():
    """Run the pipeline with Optuna optimization."""
    import sys
    sys.path.insert(0, "src")

    # Generate test data with sparse product
    print("=" * 60)
    print("STEP 1: Generating test data (with sparse product)")
    print("=" * 60)
    data_path = generate_test_data_with_sparse_product()

    # Create Optuna config
    print("\n" + "=" * 60)
    print("STEP 2: Creating Optuna-enabled configuration")
    print("=" * 60)
    config_dir = create_optuna_config()

    # Run the pipeline
    print("\n" + "=" * 60)
    print("STEP 3: Running pipeline with Optuna optimization")
    print("=" * 60)
    print("(This may take longer due to hyperparameter search)")

    from forecast.pipeline import run_pipeline

    results = run_pipeline(
        config_dir=config_dir,
        input_override=str(data_path),
        log_level="INFO",
    )

    # Results summary
    print("\n" + "=" * 60)
    print("STEP 4: Results Summary")
    print("=" * 60)

    if "error" in results:
        print(f"Pipeline failed: {results['error']}")
        return

    print(f"\nRun timestamp: {results.get('timestamp', 'N/A')}")
    print(f"Categories processed: {results.get('categories_processed', 0)}")
    print("(Product D should be skipped due to insufficient data)")

    model_results = results.get("model_results", {})
    for category, models in model_results.items():
        print(f"\n{category}:")
        for model_name, metrics in models.items():
            print(f"  {model_name}:")
            print(f"    RMSE: {metrics['rmse']:.2f}")
            print(f"    MAPE: {metrics['mape']:.2%}")
            print(f"    MAE: {metrics['mae']:.2f}")

    ensemble_results = results.get("ensemble_results", {})
    print("\nEnsemble Forecasts:")
    for category, info in ensemble_results.items():
        models_used = info.get("models_used", [])
        weights = info.get("weights", {})
        print(f"  {category}: {', '.join(models_used)}")
        if weights:
            weight_str = ", ".join(f"{k}={v:.2f}" for k, v in weights.items())
            print(f"    Weights: {weight_str}")

    # Check if Product D was skipped
    if "Product D (Sparse)" not in model_results:
        print("\n*** Product D (Sparse) was correctly skipped due to insufficient data ***")
    else:
        print("\nWARNING: Product D should have been skipped but was processed!")

    print("\n" + "=" * 60)
    print("Optuna test completed!")
    print("=" * 60)

    timestamp = results.get("timestamp", "")
    print(f"\nOutput files (timestamped with {timestamp}):")
    print(f"  - output/results/{timestamp}/model_results.xlsx")
    print(f"  - output/results/{timestamp}/ensemble_forecasts.xlsx")
    print(f"  - output/plots/{timestamp}/preprocessing/")
    print(f"  - output/plots/{timestamp}/evaluation/")
    print(f"  - output/plots/{timestamp}/forecast/")
    print(f"  - output/models/{timestamp}/")


if __name__ == "__main__":
    run_optuna_test()
